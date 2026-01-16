"""
Testes para funcionalidade de exportação Excel.

Testa a geração de arquivos Excel, formatação, conteúdo e endpoint da API.
"""

from decimal import Decimal
from io import BytesIO
from django.test import TestCase
from django.urls import reverse
from rest_framework.test import APITestCase
from rest_framework import status as http_status
from openpyxl import load_workbook

from api.models import Provider, Payroll, PayrollStatus
from services.excel_service import ExcelService


class ExcelServiceTestCase(TestCase):
    """Testes para o ExcelService."""
    
    def setUp(self):
        """Configuração inicial para os testes."""
        # Criar prestador de teste
        self.provider = Provider.objects.create(
            name="João Silva Test",
            role="Desenvolvedor",
            monthly_value=Decimal('2200.00'),
            monthly_hours=220,
            advance_percentage=Decimal('40.00'),
            vt_value=Decimal('202.40')
        )
        
        # Criar folha de teste
        self.payroll = Payroll.objects.create(
            provider=self.provider,
            reference_month="01/2026",
            base_value=Decimal('2200.00'),
            advance_value=Decimal('880.00'),
            overtime_hours_50=Decimal('10.00'),
            holiday_hours=Decimal('8.00'),
            night_hours=Decimal('20.00'),
            late_minutes=30,
            absence_hours=Decimal('8.00'),
            vt_discount=Decimal('202.40'),
            status=PayrollStatus.DRAFT
        )
    
    def test_format_currency(self):
        """Testa formatação de moeda brasileira."""
        value = Decimal('1234.56')
        formatted = ExcelService._format_currency(value)
        
        self.assertEqual(formatted, "R$ 1.234,56")
    
    def test_format_currency_zero(self):
        """Testa formatação de valor zero."""
        value = Decimal('0.00')
        formatted = ExcelService._format_currency(value)
        
        self.assertEqual(formatted, "R$ 0,00")
    
    def test_format_currency_none(self):
        """Testa formatação de valor None."""
        formatted = ExcelService._format_currency(None)
        
        self.assertEqual(formatted, "R$ 0,00")
    
    def test_get_filename(self):
        """Testa geração de nome de arquivo."""
        filename = ExcelService.get_filename(self.payroll)
        
        self.assertIn('folha_', filename)
        self.assertIn('joao_silva_test', filename)
        self.assertIn('01_2026', filename)
        self.assertTrue(filename.endswith('.xlsx'))
    
    def test_generate_excel_file(self):
        """Testa se o arquivo Excel é gerado corretamente."""
        service = ExcelService()
        excel_file = service.generate_payroll_excel(self.payroll)
        
        # Verificar que retorna BytesIO
        self.assertIsInstance(excel_file, BytesIO)
        
        # Verificar que o arquivo não está vazio
        self.assertGreater(len(excel_file.getvalue()), 0)
    
    def test_excel_content(self):
        """Testa conteúdo básico do arquivo Excel."""
        service = ExcelService()
        excel_file = service.generate_payroll_excel(self.payroll)
        
        # Carregar workbook para verificar conteúdo
        wb = load_workbook(excel_file)
        ws = wb.active
        
        # Verificar título
        self.assertEqual(ws['A1'].value, "FOLHA DE PAGAMENTO")
        
        # Verificar mês de referência
        self.assertIn("01/2026", ws['A2'].value)
        
        # Verificar nome do prestador (está em alguma célula)
        found_provider = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "João Silva Test":
                    found_provider = True
                    break
            if found_provider:
                break
        
        self.assertTrue(found_provider, "Nome do prestador não encontrado no Excel")
    
    def test_excel_values_accuracy(self):
        """Testa se os valores no Excel estão corretos."""
        service = ExcelService()
        excel_file = service.generate_payroll_excel(self.payroll)
        
        wb = load_workbook(excel_file)
        ws = wb.active
        
        # Buscar valor líquido no Excel
        found_net_value = False
        expected_net = ExcelService._format_currency(self.payroll.net_value)
        
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and expected_net in cell.value:
                    found_net_value = True
                    break
            if found_net_value:
                break
        
        self.assertTrue(found_net_value, f"Valor líquido {expected_net} não encontrado no Excel")


class ExcelExportAPITestCase(APITestCase):
    """Testes para o endpoint de exportação Excel via API."""
    
    def setUp(self):
        """Configuração inicial para os testes da API."""
        # Criar prestador de teste
        self.provider = Provider.objects.create(
            name="Maria Santos",
            role="Designer",
            monthly_value=Decimal('3000.00'),
            monthly_hours=220,
            advance_percentage=Decimal('40.00'),
            vt_value=Decimal('150.00')
        )
        
        # Criar folha de teste
        self.payroll = Payroll.objects.create(
            provider=self.provider,
            reference_month="02/2026",
            base_value=Decimal('3000.00'),
            advance_value=Decimal('1200.00'),
            overtime_hours_50=Decimal('5.00'),
            holiday_hours=Decimal('0.00'),
            night_hours=Decimal('10.00'),
            late_minutes=0,
            absence_hours=Decimal('0.00'),
            vt_discount=Decimal('150.00'),
            status=PayrollStatus.CLOSED
        )
    
    def test_export_excel_endpoint_exists(self):
        """Testa se o endpoint de exportação existe."""
        url = reverse('payroll-export-excel', kwargs={'pk': self.payroll.pk})
        response = self.client.get(url)
        
        # Não deve retornar 404
        self.assertNotEqual(response.status_code, http_status.HTTP_404_NOT_FOUND)
    
    def test_export_excel_success(self):
        """Testa exportação bem-sucedida de Excel."""
        url = reverse('payroll-export-excel', kwargs={'pk': self.payroll.pk})
        response = self.client.get(url)
        
        self.assertEqual(response.status_code, http_status.HTTP_200_OK)
        
        # Verificar content-type
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Verificar header de download
        self.assertIn('attachment', response['Content-Disposition'])
        self.assertIn('.xlsx', response['Content-Disposition'])
    
    def test_export_excel_nonexistent_payroll(self):
        """Testa tentativa de exportar folha inexistente."""
        url = reverse('payroll-export-excel', kwargs={'pk': 99999})
        response = self.client.get(url)
        
        self.assertEqual(response.status_code, http_status.HTTP_404_NOT_FOUND)
    
    def test_export_excel_file_is_valid(self):
        """Testa se o arquivo Excel retornado é válido."""
        url = reverse('payroll-export-excel', kwargs={'pk': self.payroll.pk})
        response = self.client.get(url)
        
        self.assertEqual(response.status_code, http_status.HTTP_200_OK)
        
        # Tentar carregar o arquivo com openpyxl
        try:
            excel_content = BytesIO(response.content)
            wb = load_workbook(excel_content)
            ws = wb.active
            
            # Verificar que tem conteúdo
            self.assertIsNotNone(ws['A1'].value)
            
        except Exception as e:
            self.fail(f"Arquivo Excel inválido: {str(e)}")
    
    def test_export_excel_filename(self):
        """Testa se o nome do arquivo está correto."""
        url = reverse('payroll-export-excel', kwargs={'pk': self.payroll.pk})
        response = self.client.get(url)
        
        self.assertEqual(response.status_code, http_status.HTTP_200_OK)
        
        # Verificar filename no header
        content_disposition = response['Content-Disposition']
        
        self.assertIn('folha_', content_disposition)
        self.assertIn('02_2026', content_disposition)
