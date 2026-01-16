"""
Serviço de geração de arquivos Excel para folhas de pagamento.

Este módulo utiliza openpyxl para criar arquivos Excel formatados
contendo todas as informações detalhadas da folha de pagamento.
"""

from decimal import Decimal
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


class ExcelService:
    """Serviço para geração de arquivos Excel de folhas de pagamento."""
    
    # Cores para formatação
    COLOR_HEADER = "4472C4"  # Azul
    COLOR_SECTION = "D9E1F2"  # Azul claro
    COLOR_TOTAL = "FFC000"   # Laranja
    COLOR_FINAL = "70AD47"   # Verde
    
    def __init__(self):
        self.wb = None
        self.ws = None
    
    @staticmethod
    def _format_currency(value):
        """
        Formata valor Decimal para string em formato brasileiro.
        
        Args:
            value: Valor Decimal
            
        Returns:
            String formatada (ex: "R$ 1.234,56")
        """
        if value is None:
            return "R$ 0,00"
        
        # Converter para float para formatação
        float_value = float(value)
        
        # Formatar com separadores brasileiros
        formatted = f"R$ {float_value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        
        return formatted
    
    def _apply_header_style(self, cell, font_size=12, bg_color=None):
        """
        Aplica estilo de cabeçalho à célula.
        
        Args:
            cell: Célula do openpyxl
            font_size: Tamanho da fonte
            bg_color: Cor de fundo (hex sem #)
        """
        cell.font = Font(bold=True, size=font_size)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        
        if bg_color:
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
    
    def _apply_value_style(self, cell, bold=False, align='right'):
        """
        Aplica estilo a células de valor.
        
        Args:
            cell: Célula do openpyxl
            bold: Se o texto deve ser negrito
            align: Alinhamento ('left', 'right', 'center')
        """
        if bold:
            cell.font = Font(bold=True)
        
        cell.alignment = Alignment(horizontal=align, vertical='center')
    
    def _apply_border(self, cell, style='thin'):
        """
        Aplica borda à célula.
        
        Args:
            cell: Célula do openpyxl
            style: Estilo da borda ('thin', 'medium', 'thick')
        """
        side = Side(style=style)
        cell.border = Border(left=side, right=side, top=side, bottom=side)
    
    def _apply_top_border(self, cell, style='thin'):
        """Aplica apenas borda superior à célula."""
        side = Side(style=style)
        cell.border = Border(top=side)
    
    def generate_payroll_excel(self, payroll):
        """
        Gera um arquivo Excel completo para a folha de pagamento.
        
        Args:
            payroll: Instância do modelo Payroll
            
        Returns:
            BytesIO: Arquivo Excel em memória pronto para download
        """
        # Criar workbook e worksheet
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Folha de Pagamento"
        
        # Configurar largura das colunas
        self.ws.column_dimensions['A'].width = 40
        self.ws.column_dimensions['B'].width = 20
        
        current_row = 1
        
        # ============================================================
        # CABEÇALHO
        # ============================================================
        
        # Título principal
        title_cell = self.ws.cell(row=current_row, column=1, value="FOLHA DE PAGAMENTO")
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        self.ws.merge_cells(f'A{current_row}:B{current_row}')
        current_row += 1
        
        # Mês de referência
        month_cell = self.ws.cell(row=current_row, column=1, value=f"Mês de Referência: {payroll.reference_month}")
        month_cell.font = Font(bold=True, size=12)
        self.ws.merge_cells(f'A{current_row}:B{current_row}')
        current_row += 2
        
        # ============================================================
        # DADOS DO PRESTADOR
        # ============================================================
        
        provider_header = self.ws.cell(row=current_row, column=1, value="PRESTADOR")
        self._apply_header_style(provider_header, font_size=12, bg_color=self.COLOR_SECTION)
        self.ws.merge_cells(f'A{current_row}:B{current_row}')
        current_row += 1
        
        # Nome
        self.ws.cell(row=current_row, column=1, value="Nome:")
        name_cell = self.ws.cell(row=current_row, column=2, value=payroll.provider.name)
        name_cell.font = Font(bold=True)
        current_row += 1
        
        # Função
        self.ws.cell(row=current_row, column=1, value="Função:")
        self.ws.cell(row=current_row, column=2, value=payroll.provider.role)
        current_row += 1
        
        # Salário base
        self.ws.cell(row=current_row, column=1, value="Salário Base Mensal:")
        base_cell = self.ws.cell(row=current_row, column=2, value=self._format_currency(payroll.base_value))
        self._apply_value_style(base_cell, bold=True)
        current_row += 1
        
        # Valor/hora
        self.ws.cell(row=current_row, column=1, value="Valor por Hora:")
        hourly_cell = self.ws.cell(row=current_row, column=2, value=self._format_currency(payroll.hourly_rate))
        self._apply_value_style(hourly_cell)
        current_row += 2
        
        # ============================================================
        # PROVENTOS
        # ============================================================
        
        proventos_header = self.ws.cell(row=current_row, column=1, value="PROVENTOS")
        self._apply_header_style(proventos_header, font_size=12, bg_color=self.COLOR_HEADER)
        valor_header = self.ws.cell(row=current_row, column=2, value="Valor")
        self._apply_header_style(valor_header, font_size=12, bg_color=self.COLOR_HEADER)
        valor_header.alignment = Alignment(horizontal='right', vertical='center')
        current_row += 1
        
        # Lista de proventos
        proventos = [
            ("Salário base (após adiantamento)", payroll.remaining_value),
        ]
        
        if payroll.overtime_hours_50 > 0:
            proventos.append((
                f"Horas extras 50% ({payroll.overtime_hours_50}h)",
                payroll.overtime_amount
            ))
        
        if payroll.holiday_hours > 0:
            proventos.append((
                f"Feriados trabalhados ({payroll.holiday_hours}h)",
                payroll.holiday_amount
            ))
        
        if payroll.dsr_amount > 0:
            proventos.append(("DSR (Descanso Semanal Remunerado)", payroll.dsr_amount))
        
        if payroll.night_hours > 0:
            proventos.append((
                f"Adicional noturno ({payroll.night_hours}h)",
                payroll.night_shift_amount
            ))
        
        # Adicionar proventos à planilha
        for desc, valor in proventos:
            self.ws.cell(row=current_row, column=1, value=desc)
            value_cell = self.ws.cell(row=current_row, column=2, value=self._format_currency(valor))
            self._apply_value_style(value_cell)
            current_row += 1
        
        # Total de proventos
        self._apply_top_border(self.ws.cell(row=current_row, column=1), style='medium')
        self._apply_top_border(self.ws.cell(row=current_row, column=2), style='medium')
        
        total_prov_desc = self.ws.cell(row=current_row, column=1, value="TOTAL PROVENTOS")
        total_prov_desc.font = Font(bold=True)
        total_prov_desc.fill = PatternFill(start_color=self.COLOR_TOTAL, end_color=self.COLOR_TOTAL, fill_type='solid')
        
        total_prov_value = self.ws.cell(row=current_row, column=2, value=self._format_currency(payroll.total_earnings))
        total_prov_value.font = Font(bold=True)
        total_prov_value.alignment = Alignment(horizontal='right', vertical='center')
        total_prov_value.fill = PatternFill(start_color=self.COLOR_TOTAL, end_color=self.COLOR_TOTAL, fill_type='solid')
        current_row += 2
        
        # ============================================================
        # DESCONTOS
        # ============================================================
        
        descontos_header = self.ws.cell(row=current_row, column=1, value="DESCONTOS")
        self._apply_header_style(descontos_header, font_size=12, bg_color=self.COLOR_HEADER)
        valor_header2 = self.ws.cell(row=current_row, column=2, value="Valor")
        self._apply_header_style(valor_header2, font_size=12, bg_color=self.COLOR_HEADER)
        valor_header2.alignment = Alignment(horizontal='right', vertical='center')
        current_row += 1
        
        # Lista de descontos
        descontos = [
            (f"Adiantamento quinzenal ({float(payroll.advance_value / payroll.base_value * 100):.0f}%)", payroll.advance_value),
        ]
        
        if payroll.late_minutes > 0:
            descontos.append((
                f"Atrasos ({payroll.late_minutes} minutos)",
                payroll.late_discount
            ))
        
        if payroll.absence_hours > 0:
            descontos.append((
                f"Faltas ({payroll.absence_hours}h)",
                payroll.absence_discount
            ))
        
        if payroll.vt_discount > 0:
            descontos.append(("Vale transporte", payroll.vt_discount))
        
        if payroll.manual_discounts > 0:
            descontos.append(("Descontos manuais", payroll.manual_discounts))
        
        # Adicionar descontos à planilha
        for desc, valor in descontos:
            self.ws.cell(row=current_row, column=1, value=desc)
            value_cell = self.ws.cell(row=current_row, column=2, value=self._format_currency(valor))
            self._apply_value_style(value_cell)
            current_row += 1
        
        # Total de descontos
        self._apply_top_border(self.ws.cell(row=current_row, column=1), style='medium')
        self._apply_top_border(self.ws.cell(row=current_row, column=2), style='medium')
        
        total_desc_desc = self.ws.cell(row=current_row, column=1, value="TOTAL DESCONTOS")
        total_desc_desc.font = Font(bold=True)
        total_desc_desc.fill = PatternFill(start_color=self.COLOR_TOTAL, end_color=self.COLOR_TOTAL, fill_type='solid')
        
        total_desc_value = self.ws.cell(row=current_row, column=2, value=self._format_currency(payroll.total_discounts))
        total_desc_value.font = Font(bold=True)
        total_desc_value.alignment = Alignment(horizontal='right', vertical='center')
        total_desc_value.fill = PatternFill(start_color=self.COLOR_TOTAL, end_color=self.COLOR_TOTAL, fill_type='solid')
        current_row += 2
        
        # ============================================================
        # VALOR LÍQUIDO FINAL
        # ============================================================
        
        # Adicionar linha de separação visual
        sep_cell1 = self.ws.cell(row=current_row, column=1, value="═" * 50)
        sep_cell2 = self.ws.cell(row=current_row, column=2, value="")
        self.ws.merge_cells(f'A{current_row}:B{current_row}')
        sep_cell1.font = Font(bold=True)
        current_row += 1
        
        # Valor líquido
        liquido_desc = self.ws.cell(row=current_row, column=1, value="VALOR LÍQUIDO A PAGAR")
        liquido_desc.font = Font(bold=True, size=14)
        liquido_desc.fill = PatternFill(start_color=self.COLOR_FINAL, end_color=self.COLOR_FINAL, fill_type='solid')
        
        liquido_value = self.ws.cell(row=current_row, column=2, value=self._format_currency(payroll.net_value))
        liquido_value.font = Font(bold=True, size=14)
        liquido_value.alignment = Alignment(horizontal='right', vertical='center')
        liquido_value.fill = PatternFill(start_color=self.COLOR_FINAL, end_color=self.COLOR_FINAL, fill_type='solid')
        current_row += 1
        
        # Linha de separação final
        sep_cell3 = self.ws.cell(row=current_row, column=1, value="═" * 50)
        self.ws.merge_cells(f'A{current_row}:B{current_row}')
        sep_cell3.font = Font(bold=True)
        current_row += 2
        
        # ============================================================
        # OBSERVAÇÕES E NOTAS
        # ============================================================
        
        # Nota sobre adiantamento
        if payroll.advance_value > 0:
            nota_cell = self.ws.cell(
                row=current_row, 
                column=1, 
                value=f"Adiantamento de {self._format_currency(payroll.advance_value)} já foi pago anteriormente."
            )
            nota_cell.font = Font(italic=True, size=10)
            self.ws.merge_cells(f'A{current_row}:B{current_row}')
            current_row += 1
        
        # Observações adicionais
        if payroll.notes:
            current_row += 1
            obs_header = self.ws.cell(row=current_row, column=1, value="Observações:")
            obs_header.font = Font(bold=True)
            current_row += 1
            
            obs_cell = self.ws.cell(row=current_row, column=1, value=payroll.notes)
            obs_cell.alignment = Alignment(wrap_text=True)
            self.ws.merge_cells(f'A{current_row}:B{current_row}')
            current_row += 1
        
        # Rodapé
        current_row += 2
        footer = self.ws.cell(
            row=current_row, 
            column=1, 
            value="Documento gerado automaticamente pelo Sistema de Folha de Pagamento PJ"
        )
        footer.font = Font(italic=True, size=9, color="666666")
        footer.alignment = Alignment(horizontal='center')
        self.ws.merge_cells(f'A{current_row}:B{current_row}')
        
        # Salvar em BytesIO
        excel_file = BytesIO()
        self.wb.save(excel_file)
        excel_file.seek(0)
        
        return excel_file
    
    @staticmethod
    def get_filename(payroll):
        """
        Gera nome do arquivo Excel.
        
        Args:
            payroll: Instância do modelo Payroll
            
        Returns:
            String com nome do arquivo (sem espaços, caracteres especiais)
        """
        # Normalizar nome do prestador (remover acentos e espaços)
        import unicodedata
        
        provider_name = payroll.provider.name.lower()
        provider_name = unicodedata.normalize('NFKD', provider_name)
        provider_name = provider_name.encode('ascii', 'ignore').decode('ascii')
        provider_name = provider_name.replace(' ', '_')
        
        # Normalizar mês (01/2026 -> 01_2026)
        month = payroll.reference_month.replace('/', '_')
        
        return f"folha_{provider_name}_{month}.xlsx"
